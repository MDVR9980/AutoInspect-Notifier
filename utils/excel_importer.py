# utils/excel_importer.py

"""
ماژول وارد کردن داده‌ها از فایل‌های اکسل
"""

from pathlib import Path
from typing import List, Tuple, Optional

import openpyxl


class ExcelImporter:
    """کلاس مدیریت ورود داده از فایل اکسل"""

    def __init__(self):
        self.supported_extensions = [".xlsx", ".xls"]

    def read_excel(
        self,
        file_path: str,
        sheet_name: Optional[str] = None
    ) -> Tuple[bool, List[dict], str]:
        """
        خواندن فایل اکسل

        Returns:
            (success, data, message)
        """

        try:
            path = Path(file_path)

            # بررسی وجود فایل
            if not path.exists():
                return False, [], "فایل یافت نشد"

            # بررسی فرمت فایل
            if path.suffix.lower() not in self.supported_extensions:
                return (
                    False,
                    [],
                    f"فرمت فایل پشتیبانی نمی‌شود: {path.suffix}"
                )

            # بارگذاری فایل
            workbook = openpyxl.load_workbook(
                file_path,
                data_only=True
            )

            # انتخاب شیت
            if sheet_name:
                if sheet_name not in workbook.sheetnames:
                    return (
                        False,
                        [],
                        f"شیت {sheet_name} یافت نشد"
                    )

                sheet = workbook[sheet_name]
            else:
                sheet = workbook.active

            rows = list(sheet.iter_rows(values_only=True))

            if not rows:
                workbook.close()
                return False, [], "فایل اکسل خالی است"

            # هدرها
            headers = []
            for index, header in enumerate(rows[0], start=1):
                if header:
                    headers.append(str(header).strip())
                else:
                    headers.append(f"column_{index}")

            # داده‌ها
            data = []

            for row in rows[1:]:

                # رد کردن سطر خالی
                if not any(row):
                    continue

                item = {}

                for header, value in zip(headers, row):

                    if value is None:
                        item[header] = ""
                    else:
                        item[header] = str(value).strip()

                data.append(item)

            workbook.close()

            if not data:
                return False, [], "هیچ داده‌ای پیدا نشد"

            return (
                True,
                data,
                f"{len(data)} رکورد خوانده شد"
            )

        except PermissionError:
            return (
                False,
                [],
                "فایل در حال استفاده است، لطفاً آن را ببندید"
            )

        except Exception as e:
            return (
                False,
                [],
                f"خطا در خواندن فایل اکسل: {str(e)}"
            )

    def extract_phone_and_plate(
        self,
        data: List[dict]
    ) -> List[Tuple[str, str]]:
        """
        استخراج شماره موبایل و پلاک
        """

        results = []

        phone_keys = [
            "شماره موبایل",
            "موبایل",
            "شماره",
            "mobile",
            "phone",
            "Phone",
            "Mobile"
        ]

        plate_keys = [
            "پلاک",
            "شماره پلاک",
            "plate",
            "Plate",
            "license_plate"
        ]

        for row in data:

            phone = ""
            plate = ""

            # پیدا کردن شماره موبایل
            for key in phone_keys:
                if key in row and row[key]:
                    phone = self._clean_phone(row[key])
                    break

            # پیدا کردن پلاک
            for key in plate_keys:
                if key in row and row[key]:
                    plate = self._clean_plate(row[key])
                    break

            if phone and plate:
                results.append((phone, plate))

        return results

    def validate_phone(self, phone: str) -> bool:
        """
        اعتبارسنجی شماره موبایل
        """

        if not phone:
            return False

        phone = self._clean_phone(phone)

        if len(phone) != 11:
            return False

        if not phone.startswith("09"):
            return False

        return phone.isdigit()

    def validate_plate(self, plate: str) -> bool:
        """
        اعتبارسنجی پلاک
        """

        if not plate:
            return False

        plate = self._clean_plate(plate)

        return len(plate) >= 5

    def process_excel_data(
        self,
        data: List[dict]
    ) -> Tuple[List[dict], List[dict]]:
        """
        پردازش و اعتبارسنجی داده‌های اکسل

        Returns:
            (valid_data, invalid_data)
        """

        valid_data = []
        invalid_data = []

        extracted = self.extract_phone_and_plate(data)

        for phone, plate in extracted:

            item = {
                "phone": phone,
                "plate": plate
            }

            errors = []

            if not self.validate_phone(phone):
                errors.append("شماره موبایل نامعتبر")

            if not self.validate_plate(plate):
                errors.append("پلاک نامعتبر")

            if errors:
                item["errors"] = errors
                invalid_data.append(item)
            else:
                valid_data.append(item)

        return valid_data, invalid_data

    def get_sheet_names(self, file_path: str) -> Tuple[bool, List[str], str]:
        """
        دریافت نام شیت‌ها
        """

        try:
            workbook = openpyxl.load_workbook(
                file_path,
                read_only=True
            )

            sheet_names = workbook.sheetnames

            workbook.close()

            return True, sheet_names, "لیست شیت‌ها دریافت شد"

        except Exception as e:
            return (
                False,
                [],
                f"خطا در خواندن شیت‌ها: {str(e)}"
            )

    def preview_data(
        self,
        file_path: str,
        rows_count: int = 5
    ) -> Tuple[bool, List[dict], str]:
        """
        پیش‌نمایش چند سطر اول فایل
        """

        success, data, message = self.read_excel(file_path)

        if not success:
            return False, [], message

        preview = data[:rows_count]

        return (
            True,
            preview,
            f"{len(preview)} سطر برای پیش‌نمایش آماده شد"
        )

    def _clean_phone(self, phone: str) -> str:
        """
        پاکسازی شماره موبایل
        """

        phone = str(phone).strip()

        replacements = [
            " ",
            "-",
            "(",
            ")",
            "+98",
            ".0"
        ]

        for item in replacements:
            phone = phone.replace(item, "")

        if phone.startswith("98") and len(phone) == 12:
            phone = "0" + phone[2:]

        return phone

    def _clean_plate(self, plate: str) -> str:
        """
        پاکسازی پلاک
        """

        plate = str(plate).strip()

        replacements = [
            "-",
            "_",
            "  "
        ]

        for item in replacements:
            plate = plate.replace(item, " ")

        return " ".join(plate.split())
